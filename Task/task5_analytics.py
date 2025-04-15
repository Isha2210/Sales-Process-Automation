"""
Task 5: Campaign Analytics and Reporting
---------------------------------------
This script analyzes marketing campaign data, generates reports,
and provides visualization of key metrics.
"""

import os
import json
import logging
import datetime
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger()


class CampaignAnalytics:
    def __init__(self, campaign_id=None, data_path=None):
        """
        Initialize the analytics engine with campaign data.

        Args:
            campaign_id (str): The ID of the campaign to analyze
            data_path (str): Directory containing campaign data files
        """
        self.campaign_id = campaign_id
        self.data_path = data_path
        self.tracking_data = []
        self.lead_data = []
        self.output_dir = "analytics_output"

        # Create output directory if it doesn't exist
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

        logger.info(f"Analytics initialized for campaign: {campaign_id}")

    def load_data(self):
        """Load tracking and lead data for the specified campaign."""
        try:
            # Load tracking data
            tracking_file = os.path.join(self.data_path, f"campaign_data_{self.campaign_id}.json")
            if os.path.exists(tracking_file):
                with open(tracking_file, 'r') as f:
                    self.tracking_data = json.load(f)
                logger.info(f"Loaded {len(self.tracking_data)} tracking records from {os.path.basename(tracking_file)}")
            else:
                logger.warning(f"Tracking data file not found: {tracking_file}")

            # Load lead data if available
            lead_file = os.path.join(self.data_path, f"leads_{self.campaign_id}.json")
            if os.path.exists(lead_file):
                with open(lead_file, 'r') as f:
                    self.lead_data = json.load(f)
                logger.info(f"Loaded {len(self.lead_data)} lead records")
            else:
                logger.info("No separate lead data file found. Using tracking data for lead analysis.")
                # Extract lead data from tracking data if no specific lead file
                self.lead_data = [record for record in self.tracking_data if record.get('converted_to_lead', False)]

            return True

        except Exception as e:
            logger.error(f"Error loading data: {str(e)}")
            return False

    def generate_engagement_metrics(self):
        """Calculate engagement metrics from tracking data."""
        if not self.tracking_data:
            return {}

        total_visits = len(self.tracking_data)
        unique_visitors = len(set(item.get('visitor_id', item.get('id', 'unknown')) for item in self.tracking_data))

        # Count conversions
        conversions = sum(1 for item in self.tracking_data if item.get('converted_to_lead', False))

        # Calculate engagement time statistics
        engagement_times = [item.get('engagement_time', 0) for item in self.tracking_data]
        avg_engagement = np.mean(engagement_times) if engagement_times else 0
        max_engagement = np.max(engagement_times) if engagement_times else 0

        # Traffic sources
        sources = {}
        for item in self.tracking_data:
            source = item.get('source', 'unknown')
            sources[source] = sources.get(source, 0) + 1

        # Calculate conversion rate
        conversion_rate = (conversions / total_visits) * 100 if total_visits > 0 else 0

        return {
            "total_visits": total_visits,
            "unique_visitors": unique_visitors,
            "conversions": conversions,
            "conversion_rate": conversion_rate,
            "avg_engagement_time": float(avg_engagement),  # Convert numpy type to native Python
            "max_engagement_time": float(max_engagement),  # Convert numpy type to native Python
            "traffic_sources": sources
        }

    def analyze_leads(self):
        """Analyze lead quality and demographics."""
        if not self.lead_data:
            return {}

        # Extract demographic data
        demographics = {}
        lead_scores = []
        lead_statuses = {}

        for lead in self.lead_data:
            # Process demographics
            if 'demographics' in lead:
                for key, value in lead['demographics'].items():
                    if key not in demographics:
                        demographics[key] = {}
                    demographics[key][value] = demographics[key].get(value, 0) + 1

            # Process lead scores
            score = lead.get('lead_score', 0)
            lead_scores.append(score)

            # Process lead statuses
            status = lead.get('status', 'unknown')
            lead_statuses[status] = lead_statuses.get(status, 0) + 1

        # Calculate lead score metrics
        avg_lead_score = np.mean(lead_scores) if lead_scores else 0
        max_lead_score = np.max(lead_scores) if lead_scores else 0
        min_lead_score = np.min(lead_scores) if lead_scores else 0

        return {
            "total_leads": len(self.lead_data),
            "demographics": demographics,
            "lead_scores": {
                "average": float(avg_lead_score),  # Convert numpy type to native Python
                "max": float(max_lead_score),  # Convert numpy type to native Python
                "min": float(min_lead_score)  # Convert numpy type to native Python
            },
            "lead_statuses": lead_statuses
        }

    def generate_lead_report(self):
        """Generate Excel report with lead information."""
        if not self.lead_data:
            logger.warning("No lead data available for reporting")
            return None

        # Convert lead data to DataFrame for easier manipulation
        lead_df = pd.DataFrame(self.lead_data)

        # Extract date if available
        for lead in lead_df.iterrows():
            if 'timestamp' in lead[1]:
                try:
                    lead_df.at[lead[0], 'date'] = datetime.datetime.fromisoformat(
                        lead[1]['timestamp'].replace('Z', '+00:00')
                    ).strftime('%Y-%m-%d')
                except:
                    lead_df.at[lead[0], 'date'] = 'Unknown'

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Lead Report"

        # Add title
        ws['A1'] = f"Lead Report - Campaign {self.campaign_id}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')

        # Add report generation timestamp
        ws['A2'] = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:F2')

        # Add summary data
        row = 4
        ws[f'A{row}'] = "Lead Summary"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        ws[f'A{row}'] = "Total Leads:"
        ws[f'B{row}'] = len(self.lead_data)

        # Add lead status breakdown if available
        if 'status' in lead_df.columns:
            status_counts = lead_df['status'].value_counts()

            row += 2
            ws[f'A{row}'] = "Lead Status Breakdown"
            ws[f'A{row}'].font = Font(bold=True)
            ws.merge_cells(f'A{row}:F{row}')

            row += 1
            for status, count in status_counts.items():
                ws[f'A{row}'] = status
                ws[f'B{row}'] = count
                row += 1

        # Add lead data table
        row += 2
        ws[f'A{row}'] = "Lead Details"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        # Select and reorder columns for the report
        columns_to_include = [
            'id', 'visitor_id', 'lead_score', 'status', 'date',
            'source', 'email', 'phone', 'name'
        ]

        display_columns = [col for col in columns_to_include if col in lead_df.columns]
        lead_table = lead_df[display_columns].copy() if display_columns else lead_df

        # Add data table
        for r_idx, row_data in enumerate(dataframe_to_rows(lead_table, index=False, header=True)):
            for c_idx, value in enumerate(row_data):
                cell = ws.cell(row=row + r_idx, column=c_idx + 1, value=value)
                if r_idx == 0:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Save the workbook
        filename = os.path.join(self.output_dir, f"lead_report_{self.campaign_id}.xlsx")
        wb.save(filename)
        logger.info(f"Data saved to {filename}")

        return filename

    def generate_visualizations(self):
        """Generate visualization charts for campaign metrics."""
        # Skip if no data
        if not self.tracking_data:
            return None

        # Get metrics data
        metrics = self.generate_engagement_metrics()
        lead_metrics = self.analyze_leads()

        # Create figure for multiple plots
        fig, axs = plt.subplots(2, 2, figsize=(15, 12))

        # Plot 1: Traffic Sources Pie Chart
        sources = metrics.get('traffic_sources', {})
        if sources:
            ax = axs[0, 0]
            labels = list(sources.keys())
            sizes = list(sources.values())
            ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
            ax.set_title('Traffic Sources')
            ax.axis('equal')

        # Plot 2: Engagement Time Histogram
        if self.tracking_data:
            ax = axs[0, 1]
            engagement_times = [item.get('engagement_time', 0) for item in self.tracking_data]
            ax.hist(engagement_times, bins=10, edgecolor='black')
            ax.set_title('Engagement Time Distribution')
            ax.set_xlabel('Engagement Time (seconds)')
            ax.set_ylabel('Number of Visitors')

        # Plot 3: Lead Scores
        if self.lead_data:
            ax = axs[1, 0]
            lead_scores = [lead.get('lead_score', 0) for lead in self.lead_data]
            if lead_scores:
                ax.hist(lead_scores, bins=5, edgecolor='black')
                ax.set_title('Lead Score Distribution')
                ax.set_xlabel('Lead Score')
                ax.set_ylabel('Number of Leads')

        # Plot 4: Lead Status
        statuses = lead_metrics.get('lead_statuses', {})
        if statuses:
            ax = axs[1, 1]
            status_labels = list(statuses.keys())
            status_counts = list(statuses.values())
            ax.bar(status_labels, status_counts)
            ax.set_title('Lead Status Distribution')
            ax.set_xlabel('Status')
            ax.set_ylabel('Number of Leads')
            plt.setp(ax.get_xticklabels(), rotation=45, ha='right')

        # Adjust layout
        plt.tight_layout()

        # Save figure
        chart_file = os.path.join(self.output_dir, f"campaign_analytics_{self.campaign_id}.png")
        plt.savefig(chart_file)

        return chart_file

    def save_analytics_results(self):
        """Save analytics results to JSON file."""
        # Skip if no data
        if not self.tracking_data:
            return None

        # Gather all analytics data
        results = {
            "campaign_id": self.campaign_id,
            "analysis_timestamp": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "engagement_metrics": self.generate_engagement_metrics(),
            "lead_metrics": self.analyze_leads()
        }

        # Save to JSON file
        filename = os.path.join(self.output_dir, f"analytics_results_{self.campaign_id}.json")
        try:
            with open(filename, 'w') as f:
                json.dump(results, f, indent=2, cls=NumpyJSONEncoder)
            logger.info(f"Analytics results saved to {filename}")
            return filename
        except Exception as e:
            logger.error(f"Error in analytics: {str(e)}")
            return None

    def run_full_analysis(self):
        """Run complete analysis workflow and generate all outputs."""
        if not self.load_data():
            logger.error("Failed to load campaign data. Analysis aborted.")
            return False

        print("=" * 60)
        print("                 Campaign Analytics Summary                 ")
        print("=" * 60)

        # Generate all outputs
        try:
            # Get and display engagement metrics
            metrics = self.generate_engagement_metrics()
            if metrics:
                print(f"\nEngagement Summary:")
                print(f"  Total Visits: {metrics['total_visits']}")
                print(f"  Unique Visitors: {metrics['unique_visitors']}")
                print(f"  Conversions: {metrics['conversions']}")
                print(f"  Conversion Rate: {metrics['conversion_rate']:.2f}%")
                print(f"  Avg. Engagement Time: {metrics['avg_engagement_time']:.2f} seconds")

            # Get and display lead metrics
            lead_metrics = self.analyze_leads()
            if lead_metrics:
                print(f"\nLead Summary:")
                print(f"  Total Leads: {lead_metrics['total_leads']}")
                if 'lead_scores' in lead_metrics:
                    print(f"  Avg. Lead Score: {lead_metrics['lead_scores']['average']:.2f}")

            # Generate Excel report
            report_file = self.generate_lead_report()
            if report_file:
                print(f"\nLead report generated: {report_file}")

            # Generate visualizations
            chart_file = self.generate_visualizations()
            if chart_file:
                print(f"Analytics charts generated: {chart_file}")

            # Save analytics results
            results_file = self.save_analytics_results()
            if results_file:
                print(f"Analytics results saved: {results_file}")

            return True

        except Exception as e:
            logger.error(f"Error in analysis: {str(e)}")
            print(f"An error occurred: {str(e)}")
            return False


class NumpyJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder that handles NumPy data types."""

    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return super(NumpyJSONEncoder, self).default(obj)


def find_data_directory():
    """Find the campaign data directory by looking in common locations."""
    # Possible data directory locations
    possible_locations = [
        "campaign_data",  # Current directory
        "data/campaign_data",  # data subdirectory
        "../campaign_data",  # Parent directory
        "../data/campaign_data",  # Parent's data subdirectory
        os.path.join(os.path.dirname(__file__), "campaign_data")  # Script directory
    ]

    # Custom locations from your file path
    custom_locations = [
        "E:/ISHA/Task/campaign_data",  # Same directory as your script
        "E:/ISHA/campaign_data",  # One level up from script
        "E:/campaign_data",  # Root level
    ]

    # Add custom locations to the search
    possible_locations.extend(custom_locations)

    # Find the first location that exists and contains campaign data files
    for location in possible_locations:
        if os.path.exists(location):
            # Check if it contains campaign data files
            files = os.listdir(location)
            for file in files:
                if file.startswith("campaign_data_") and file.endswith(".json"):
                    print(f"Found campaign data directory: {location}")
                    return location

    # If no existing directory found, try to create one
    try:
        os.makedirs("campaign_data", exist_ok=True)
        print("Created a new campaign data directory: campaign_data")
        return "campaign_data"
    except:
        pass

    # If all else fails, use the current directory
    print("Warning: Could not find or create a campaign data directory. Using current directory.")
    return "."


def create_sample_data(data_path):
    """Create sample campaign data files for testing."""
    if not os.path.exists(data_path):
        os.makedirs(data_path)

    # Create a sample campaign ID
    campaign_id = datetime.datetime.now().strftime("%Y%m%d%H%M%S")

    # Sample tracking data
    tracking_data = [
        {
            "id": "v1001",
            "visitor_id": "visitor1",
            "timestamp": "2025-04-12T10:30:00Z",
            "source": "google",
            "engagement_time": 120,
            "converted_to_lead": True,
            "lead_score": 85,
            "status": "new",
            "email": "john.doe@example.com",
            "name": "John Doe"
        },
        {
            "id": "v1002",
            "visitor_id": "visitor2",
            "timestamp": "2025-04-12T11:15:00Z",
            "source": "facebook",
            "engagement_time": 90,
            "converted_to_lead": False
        },
        {
            "id": "v1003",
            "visitor_id": "visitor3",
            "timestamp": "2025-04-12T12:00:00Z",
            "source": "direct",
            "engagement_time": 180,
            "converted_to_lead": True,
            "lead_score": 65,
            "status": "contacted",
            "email": "jane.smith@example.com",
            "name": "Jane Smith"
        },
        {
            "id": "v1004",
            "visitor_id": "visitor4",
            "timestamp": "2025-04-12T13:45:00Z",
            "source": "linkedin",
            "engagement_time": 150,
            "converted_to_lead": True,
            "lead_score": 90,
            "status": "qualified",
            "email": "alex.brown@example.com",
            "name": "Alex Brown"
        },
        {
            "id": "v1005",
            "visitor_id": "visitor5",
            "timestamp": "2025-04-12T14:30:00Z",
            "source": "twitter",
            "engagement_time": 60,
            "converted_to_lead": False
        }
    ]

    # Save tracking data
    tracking_file = os.path.join(data_path, f"campaign_data_{campaign_id}.json")
    with open(tracking_file, 'w') as f:
        json.dump(tracking_data, f, indent=2)

    print(f"Created sample campaign data: {tracking_file}")
    return campaign_id


def list_available_campaigns(data_path):
    """List all available campaigns in the data directory."""
    campaigns = []

    if not os.path.exists(data_path):
        print(f"Data directory not found: {data_path}")
        return []

    for filename in os.listdir(data_path):
        if filename.startswith("campaign_data_") and filename.endswith(".json"):
            campaign_id = filename.replace("campaign_data_", "").replace(".json", "")
            file_path = os.path.join(data_path, filename)
            mod_time = datetime.datetime.fromtimestamp(os.path.getmtime(file_path))
            campaigns.append((campaign_id, mod_time))

    # Sort by modification time (newest first)
    campaigns.sort(key=lambda x: x[1], reverse=True)
    return campaigns


def main():
    """Main function to run the campaign analytics tool."""
    print("=" * 60)
    print("          Task 5: Campaign Analytics and Reporting          ")
    print("=" * 60)

    # Find data directory
    data_path = find_data_directory()

    # List available campaigns
    campaigns = list_available_campaigns(data_path)

    # If no campaigns found, offer to create sample data
    if not campaigns:
        print("No campaign data found.")
        create_sample = input("Would you like to create sample campaign data for testing? (y/n): ")
        if create_sample.lower() == 'y':
            campaign_id = create_sample_data(data_path)
            campaigns = [(campaign_id, datetime.datetime.now())]
        else:
            print("Exiting. Please add campaign data files before running analysis.")
            return

    # Display available campaigns
    print("\nAvailable campaigns:")
    for i, (campaign_id, mod_time) in enumerate(campaigns, 1):
        print(f"{i}. Campaign {campaign_id} (Last modified: {mod_time.strftime('%Y-%m-%d %H:%M:%S')})")

    # Get campaign selection from user
    try:
        selection = input("Enter campaign number to analyze (or press Enter for most recent): ")
        if selection.strip() == "":
            selected_campaign = campaigns[0][0]
        else:
            idx = int(selection) - 1
            if 0 <= idx < len(campaigns):
                selected_campaign = campaigns[idx][0]
            else:
                print("Invalid selection. Using most recent campaign.")
                selected_campaign = campaigns[0][0]
    except (ValueError, IndexError):
        print("Invalid input. Using most recent campaign.")
        selected_campaign = campaigns[0][0]

    print(f"Analyzing campaign: {selected_campaign}")

    # Create analytics instance and run analysis
    analytics = CampaignAnalytics(selected_campaign, data_path)
    analytics.run_full_analysis()


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"An error occurred: {str(e)}")